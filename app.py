"""starphotocard-go — Group Order management web app.

Public side (English): buyers browse active GOs and submit orders.
Admin side (password-gated): manage GO campaigns, view buyer orders, import xlsx.
On every new order: a Telegram message is pushed to the admin chat (if env vars set).
"""
from __future__ import annotations

import csv
import io
import os
import re
import secrets
import time
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any

import httpx
from fastapi import FastAPI, Request, Form, Cookie, HTTPException, UploadFile, File, Depends
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, PlainTextResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from supabase import create_client, Client

# ----------------------------------------------------------------------------
# Config
# ----------------------------------------------------------------------------

SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
APP_PASSWORD = os.environ.get("APP_PASSWORD", "")
SESSION_HOURS = int(os.environ.get("SESSION_HOURS", "24"))
# Kakao "memo to self" API — used in place of Telegram. OAuth token refreshes every ~6h.
KAKAO_REST_API_KEY = os.environ.get("KAKAO_REST_API_KEY", "")
KAKAO_REDIRECT_URI = os.environ.get("KAKAO_REDIRECT_URI", "https://starphotocard-go.up.railway.app/admin/kakao/callback")
# Optional — required only if the Kakao app has "Client Secret" enabled in 보안 settings.
KAKAO_CLIENT_SECRET = os.environ.get("KAKAO_CLIENT_SECRET", "")
# Telegram kept as fallback if Kakao not configured (gracefully no-ops if neither is set)
TG_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TG_CHAT = os.environ.get("TELEGRAM_ADMIN_CHAT_ID", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    # Allow boot without env so the user can see /health while configuring; routes that touch DB will 500.
    print("[WARN] SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY missing — DB calls will fail.")

sb: Optional[Client] = None
if SUPABASE_URL and SUPABASE_KEY:
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

app = FastAPI(title="starphotocard-go", version="0.1.0")
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# ----------------------------------------------------------------------------
# Sessions (in-memory; same pattern as ImageAutoUploader). For multi-instance
# deploys swap to Redis or Supabase, but Railway free tier runs single instance.
# ----------------------------------------------------------------------------

_sessions: Dict[str, datetime] = {}

def _now() -> datetime:
    return datetime.utcnow()

def _make_session() -> str:
    sid = secrets.token_urlsafe(32)
    _sessions[sid] = _now() + timedelta(hours=SESSION_HOURS)
    return sid

def _is_logged_in(session_id: Optional[str]) -> bool:
    if not session_id or session_id not in _sessions:
        return False
    if _sessions[session_id] < _now():
        _sessions.pop(session_id, None)
        return False
    return True

def require_admin(session_id: Optional[str] = Cookie(None)):
    """Dependency for /admin/api/* JSON endpoints — returns 401 JSON if not logged in.
    HTML pages handle auth inline by returning RedirectResponse."""
    if not _is_logged_in(session_id):
        raise HTTPException(status_code=401, detail="Login required")
    return session_id

# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------

GO_CODE_RE = re.compile(r"^(GO|ALBUM|MD|RESTOCK)-\d{3,}$")  # RESTOCK kept for backward-compat with 3 historical codes

CATEGORY_PREFIX = {"POCA_SET": "GO", "ALBUM": "ALBUM", "MD": "MD"}
# RESTOCK 카테고리는 2026-05-05 ALBUM 으로 통합. RESTOCK- prefix 의 historical 코드는 ALBUM 카테고리로 매핑됨.

def category_for_code(code: str) -> str:
    if code.startswith("GO-"): return "POCA_SET"
    if code.startswith("ALBUM-"): return "ALBUM"
    if code.startswith("MD-"): return "MD"
    if code.startswith("RESTOCK-"): return "ALBUM"  # historical: RESTOCK 통합 후 ALBUM 으로 인식
    return "POCA_SET"

def next_go_code(category: str) -> str:
    """Auto-generate the next GO code per category."""
    prefix = CATEGORY_PREFIX.get(category, "GO")
    res = sb.table("go_campaigns").select("go_code").like("go_code", f"{prefix}-%").execute()
    nums: List[int] = []
    for r in (res.data or []):
        m = re.match(rf"^{prefix}-(\d+)$", r["go_code"])
        if m: nums.append(int(m.group(1)))
    nxt = (max(nums) + 1) if nums else 1
    return f"{prefix}-{nxt:03d}"

def derived_status(deadline: Optional[date], is_closed: bool) -> str:
    if is_closed: return "closed"
    if deadline is None: return "closed"
    delta = (deadline - date.today()).days
    if delta < 0: return "closed"
    if delta <= 3: return "closing_soon"
    return "open"

def parse_date(value: Any) -> Optional[date]:
    if not value: return None
    if isinstance(value, date) and not isinstance(value, datetime): return value
    if isinstance(value, datetime): return value.date()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None

def parse_num(value: Any) -> Optional[float]:
    if value in (None, "", "X", "-"): return None
    try: return float(str(value).replace(",", "").strip())
    except ValueError: return None

async def telegram_notify(text: str) -> bool:
    if not TG_TOKEN or not TG_CHAT:
        return False
    url = f"https://api.telegram.org/bot{TG_TOKEN}/sendMessage"
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.post(url, json={"chat_id": TG_CHAT, "text": text, "disable_web_page_preview": True})
            return r.status_code == 200
    except Exception as e:
        print(f"[telegram_notify] {e}")
        return False

# ----------------------------------------------------------------------------
# Kakao "memo to self" — replaces Telegram for admin notifications.
# Flow:
#   1. Admin visits /admin/kakao/connect once → Kakao login → consent
#   2. /admin/kakao/callback receives ?code=XXX → exchanges for access_token
#      + refresh_token, persists into kakao_token (single-row).
#   3. notify_admin() reads token, refreshes if access expired (refresh_token
#      lives ~60 days; we re-issue refresh_token if Kakao returns one).
# ----------------------------------------------------------------------------

KAKAO_AUTH_URL = "https://kauth.kakao.com/oauth/authorize"
KAKAO_TOKEN_URL = "https://kauth.kakao.com/oauth/token"
KAKAO_SEND_URL = "https://kapi.kakao.com/v2/api/talk/memo/default/send"

async def _kakao_load_token() -> Optional[Dict[str, Any]]:
    if not sb: return None
    res = sb.table("kakao_token").select("*").eq("id", 1).limit(1).execute()
    rows = res.data or []
    return rows[0] if rows else None

async def _kakao_save_token(access_token: Optional[str], refresh_token: Optional[str],
                            expires_in: Optional[int], refresh_expires_in: Optional[int],
                            scope: Optional[str]):
    """Upsert into single-row kakao_token. Kakao only returns refresh_token when it's
    re-issued (when previous one is < 1 month from expiry); preserve existing one otherwise."""
    if not sb: return
    payload: Dict[str, Any] = {"id": 1, "updated_at": _now().isoformat()}
    if access_token: payload["access_token"] = access_token
    if refresh_token: payload["refresh_token"] = refresh_token
    if expires_in is not None:
        payload["access_expires_at"] = (_now() + timedelta(seconds=int(expires_in))).isoformat()
    if refresh_expires_in is not None:
        payload["refresh_expires_at"] = (_now() + timedelta(seconds=int(refresh_expires_in))).isoformat()
    if scope: payload["scope"] = scope
    sb.table("kakao_token").upsert(payload, on_conflict="id").execute()

async def _kakao_get_valid_access_token() -> Optional[str]:
    """Returns a valid access_token, refreshing via refresh_token if needed.
    Returns None if not configured or refresh_token missing/expired."""
    if not KAKAO_REST_API_KEY: return None
    row = await _kakao_load_token()
    if not row: return None
    access = row.get("access_token")
    refresh = row.get("refresh_token")
    if not refresh:
        return None  # not authorized yet — admin needs to visit /admin/kakao/connect
    # Is access still valid (with 60s buffer)?
    if access and row.get("access_expires_at"):
        try:
            exp = datetime.fromisoformat(row["access_expires_at"].replace("Z", "+00:00"))
            if exp.replace(tzinfo=None) - _now() > timedelta(seconds=60):
                return access
        except Exception:
            pass
    # Refresh
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            data = {
                "grant_type": "refresh_token",
                "client_id": KAKAO_REST_API_KEY,
                "refresh_token": refresh,
            }
            if KAKAO_CLIENT_SECRET:
                data["client_secret"] = KAKAO_CLIENT_SECRET
            r = await client.post(KAKAO_TOKEN_URL, data=data)
            j = r.json()
            if r.status_code != 200 or not j.get("access_token"):
                print(f"[kakao_refresh] failed {r.status_code}: {j}")
                return None
            await _kakao_save_token(
                access_token=j.get("access_token"),
                refresh_token=j.get("refresh_token"),  # may be absent — preserve existing
                expires_in=j.get("expires_in"),
                refresh_expires_in=j.get("refresh_token_expires_in"),
                scope=None,
            )
            return j.get("access_token")
    except Exception as e:
        print(f"[kakao_refresh] exception: {e}")
        return None

async def kakao_notify(text: str, link_url: Optional[str] = None) -> bool:
    access = await _kakao_get_valid_access_token()
    if not access: return False
    template_object = {
        "object_type": "text",
        "text": text[:4000],  # Kakao limit
        "link": {"web_url": link_url or "https://starphotocard-go.up.railway.app",
                 "mobile_web_url": link_url or "https://starphotocard-go.up.railway.app"},
        "button_title": "Open Admin",
    }
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.post(KAKAO_SEND_URL,
                                  headers={"Authorization": f"Bearer {access}"},
                                  data={"template_object": json_dumps(template_object)})
            if r.status_code == 200:
                return True
            print(f"[kakao_notify] {r.status_code}: {r.text[:200]}")
            return False
    except Exception as e:
        print(f"[kakao_notify] exception: {e}")
        return False

# Use json.dumps (avoid repeated import elsewhere)
import json as _json
def json_dumps(o: Any) -> str: return _json.dumps(o, ensure_ascii=False)

async def notify_admin(text: str, link_url: Optional[str] = None) -> bool:
    """Single entrypoint — prefers Kakao, falls back to Telegram if Kakao not configured/failed."""
    # Try Kakao first
    if KAKAO_REST_API_KEY:
        if await kakao_notify(text, link_url):
            return True
    # Fallback to Telegram
    return await telegram_notify(text)

def _enrich_status(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Add `status` and `d_day` to a list of campaign rows."""
    today = date.today()
    out = []
    for r in rows:
        dl = r.get("deadline")
        if isinstance(dl, str):
            try: dl_date = datetime.strptime(dl, "%Y-%m-%d").date()
            except ValueError: dl_date = None
        elif isinstance(dl, date):
            dl_date = dl
        else:
            dl_date = None
        status = derived_status(dl_date, bool(r.get("is_closed")))
        d_day = (dl_date - today).days if dl_date else None
        out.append({**r, "status": status, "d_day": d_day})
    return out

# ----------------------------------------------------------------------------
# Public routes
# ----------------------------------------------------------------------------

@app.get("/health")
async def health():
    return {"status": "ok", "service": "starphotocard-go", "version": "0.1.0",
            "db": bool(sb), "telegram": bool(TG_TOKEN and TG_CHAT)}

@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    by_cat: Dict[str, List[Dict[str, Any]]] = {"POCA_SET": [], "ALBUM": [], "RESTOCK": []}
    error = None
    if not sb:
        error = "Service is being configured. Please check back shortly."
        return templates.TemplateResponse("home.html", {"request": request, "by_cat": by_cat, "active_count": 0, "error": error})
    res = sb.table("go_campaigns").select("*").order("category").order("deadline", desc=False).execute()
    enriched = _enrich_status(res.data or [])
    active = [c for c in enriched if c["status"] in ("open", "closing_soon")]
    # Attach buyer_count per GO from go_campaign_aggregate view (used as social-proof badge, gated >=3 in template)
    try:
        agg = sb.table("go_campaign_aggregate").select("go_code, buyer_count").execute()
        bc_map: Dict[str, int] = {r["go_code"]: int(r["buyer_count"] or 0) for r in (agg.data or [])}
    except Exception:
        bc_map = {}
    for c in active:
        c["buyer_count"] = bc_map.get(c["go_code"], 0)
        by_cat[c["category"]].append(c)
    return templates.TemplateResponse("home.html", {"request": request, "by_cat": by_cat, "active_count": len(active), "error": None})

@app.get("/order", response_class=HTMLResponse)
async def order_form(request: Request):
    if not sb:
        return templates.TemplateResponse("order.html", {"request": request, "options": [], "error": "Service is being configured. Please check back shortly."})
    res = sb.table("go_campaigns").select("go_code, category, idol_name, album, version, retail_store, set_price_krw, set_detail").execute()
    enriched = _enrich_status(res.data or [])
    options = [c for c in enriched if c["status"] in ("open", "closing_soon")]
    return templates.TemplateResponse("order.html", {"request": request, "options": options, "error": None})

@app.post("/order")
async def order_submit(request: Request):
    if not sb:
        raise HTTPException(status_code=503, detail="Service unavailable")
    form = await request.form()
    alibaba_username = (form.get("alibaba_username") or "").strip()
    contact_email = (form.get("contact_email") or "").strip() or None
    contact_note = (form.get("contact_note") or "").strip() or None
    if not alibaba_username:
        raise HTTPException(status_code=400, detail="Alibaba username is required.")
    # Multi-row items: go_code[] + qty[]
    codes = form.getlist("go_code[]")
    qtys = form.getlist("qty[]")
    items: List[Dict[str, Any]] = []
    total_krw = 0.0
    # Pre-fetch prices
    if codes:
        price_res = sb.table("go_campaigns").select("go_code, set_price_krw").in_("go_code", codes).execute()
        price_map = {r["go_code"]: r.get("set_price_krw") for r in (price_res.data or [])}
    else:
        price_map = {}
    for code, qty in zip(codes, qtys):
        code = (code or "").strip()
        try: q = int(qty)
        except (TypeError, ValueError): continue
        if not code or q <= 0: continue
        unit = float(price_map.get(code) or 0)
        line_total = unit * q
        total_krw += line_total
        items.append({"go_code": code, "qty": q, "unit_price_krw": unit or None, "line_total_krw": line_total or None})
    if not items:
        raise HTTPException(status_code=400, detail="At least one valid order item (GO code + qty) is required.")
    # Insert order
    order_res = sb.table("go_buyer_orders").insert({
        "alibaba_username": alibaba_username,
        "contact_email": contact_email,
        "contact_note": contact_note,
        "total_krw": total_krw,
    }).execute()
    order_id = order_res.data[0]["id"]
    for it in items:
        it["order_id"] = order_id
    sb.table("go_buyer_order_items").insert(items).execute()
    # Notify admin via Telegram (fire-and-forget; doesn't block submit if Telegram is down)
    summary = "\n".join([f"  {it['go_code']} × {it['qty']}" for it in items])
    msg = (f"🎀 New GO order #{order_id}\n"
           f"Buyer: {alibaba_username}\n"
           f"Items:\n{summary}\n"
           f"Total: ₩{int(total_krw):,}")
    notified = await notify_admin(msg, link_url="https://starphotocard-go.up.railway.app/admin/orders")
    if notified:
        sb.table("go_buyer_orders").update({"notified_at": _now().isoformat()}).eq("id", order_id).execute()
    return RedirectResponse(f"/order/success?id={order_id}", status_code=302)

@app.get("/order/success", response_class=HTMLResponse)
async def order_success(request: Request, id: int):
    return templates.TemplateResponse("order_success.html", {"request": request, "order_id": id})

# ----------------------------------------------------------------------------
# Admin — login
# ----------------------------------------------------------------------------

@app.get("/admin", response_class=HTMLResponse)
async def admin_redirect(request: Request, session_id: Optional[str] = Cookie(None)):
    if _is_logged_in(session_id):
        return RedirectResponse("/admin/dashboard", status_code=302)
    return RedirectResponse("/admin/login", status_code=302)

@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request, "error": None})

@app.post("/admin/login")
async def admin_login_submit(request: Request, password: str = Form(...)):
    if not APP_PASSWORD:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Server APP_PASSWORD not configured."})
    if password != APP_PASSWORD:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Wrong password."})
    sid = _make_session()
    resp = RedirectResponse("/admin/dashboard", status_code=302)
    resp.set_cookie("session_id", sid, max_age=SESSION_HOURS * 3600, httponly=True, samesite="lax", secure=True)
    return resp

@app.get("/admin/logout")
async def admin_logout(session_id: Optional[str] = Cookie(None)):
    if session_id: _sessions.pop(session_id, None)
    resp = RedirectResponse("/admin/login", status_code=302)
    resp.delete_cookie("session_id")
    return resp

# ----------------------------------------------------------------------------
# Admin — pages (HTML, redirect to login if not authed)
# ----------------------------------------------------------------------------

def _admin_or_redirect(session_id: Optional[str]):
    if not _is_logged_in(session_id):
        return RedirectResponse("/admin/login", status_code=302)
    return None

@app.get("/admin/dashboard", response_class=HTMLResponse)
async def admin_dashboard(request: Request, session_id: Optional[str] = Cookie(None)):
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    counts = sb.table("go_campaigns").select("category", count="exact").execute()
    cnt_total = sb.table("go_campaigns").select("id", count="exact").execute().count or 0
    open_q = sb.table("go_campaigns").select("id", count="exact").eq("is_closed", False).gte("deadline", date.today().isoformat()).execute()
    cnt_open = open_q.count or 0
    orders_q = sb.table("go_buyer_orders").select("id", count="exact").execute()
    cnt_orders = orders_q.count or 0
    # Recent orders
    recent = sb.table("go_buyer_orders").select("*").order("created_at", desc=True).limit(10).execute()
    return templates.TemplateResponse("admin/dashboard.html", {
        "request": request, "cnt_total": cnt_total, "cnt_open": cnt_open, "cnt_orders": cnt_orders,
        "recent_orders": recent.data or [],
    })

@app.get("/admin/campaigns", response_class=HTMLResponse)
async def admin_campaigns(request: Request, session_id: Optional[str] = Cookie(None), q: str = "", cat: str = ""):
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    query = sb.table("go_campaigns").select("*").order("category").order("go_code")
    if cat in ("POCA_SET", "ALBUM", "RESTOCK"):
        query = query.eq("category", cat)
    res = query.execute()
    rows = _enrich_status(res.data or [])
    if q:
        ql = q.lower()
        rows = [r for r in rows if ql in (r.get("idol_name","").lower() + " " + r.get("album","").lower() + " " + r.get("retail_store","").lower())]
    return templates.TemplateResponse("admin/campaigns.html", {"request": request, "rows": rows, "q": q, "cat": cat})

@app.get("/admin/campaign/new", response_class=HTMLResponse)
async def admin_campaign_new(request: Request, session_id: Optional[str] = Cookie(None), category: str = "POCA_SET"):
    if (r := _admin_or_redirect(session_id)): return r
    return templates.TemplateResponse("admin/campaign_form.html", {
        "request": request, "row": None, "category": category, "is_new": True,
    })

@app.post("/admin/campaign/new")
async def admin_campaign_create(request: Request, session_id: Optional[str] = Cookie(None)):
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    form = await request.form()
    category = form.get("category", "POCA_SET")
    payload = _campaign_payload_from_form(form)
    payload["category"] = category
    payload["go_code"] = (form.get("go_code") or "").strip() or next_go_code(category)
    sb.table("go_campaigns").insert(payload).execute()
    return RedirectResponse("/admin/campaigns", status_code=302)

@app.get("/admin/campaign/{code}/edit", response_class=HTMLResponse)
async def admin_campaign_edit(request: Request, code: str, session_id: Optional[str] = Cookie(None)):
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    res = sb.table("go_campaigns").select("*").eq("go_code", code).single().execute()
    if not res.data: raise HTTPException(404, "Campaign not found")
    return templates.TemplateResponse("admin/campaign_form.html", {
        "request": request, "row": res.data, "category": res.data["category"], "is_new": False,
    })

@app.post("/admin/campaign/{code}/edit")
async def admin_campaign_update(request: Request, code: str, session_id: Optional[str] = Cookie(None)):
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    form = await request.form()
    payload = _campaign_payload_from_form(form)
    payload["updated_at"] = _now().isoformat()
    sb.table("go_campaigns").update(payload).eq("go_code", code).execute()
    return RedirectResponse("/admin/campaigns", status_code=302)

@app.post("/admin/campaign/{code}/delete")
async def admin_campaign_delete(request: Request, code: str, session_id: Optional[str] = Cookie(None)):
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    sb.table("go_campaigns").delete().eq("go_code", code).execute()
    return RedirectResponse("/admin/campaigns", status_code=302)

@app.post("/admin/campaign/{code}/toggle-close")
async def admin_campaign_toggle_close(request: Request, code: str, session_id: Optional[str] = Cookie(None)):
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    cur = sb.table("go_campaigns").select("is_closed").eq("go_code", code).single().execute()
    new_val = not bool(cur.data["is_closed"])
    sb.table("go_campaigns").update({"is_closed": new_val, "updated_at": _now().isoformat()}).eq("go_code", code).execute()
    return RedirectResponse("/admin/campaigns", status_code=302)

def _campaign_payload_from_form(form) -> Dict[str, Any]:
    return {
        "idol_name": (form.get("idol_name") or "").strip(),
        "album": (form.get("album") or "").strip(),
        "version": (form.get("version") or "").strip() or None,
        "retail_store": (form.get("retail_store") or "").strip() or None,
        "url": (form.get("url") or "").strip() or None,
        "set_price_krw": parse_num(form.get("set_price_krw")),
        "retail_price_krw": parse_num(form.get("retail_price_krw")),
        "discount_pct": parse_num(form.get("discount_pct")),
        "pob_store": (form.get("pob_store") or "").strip() or None,
        "pob_detail": (form.get("pob_detail") or "").strip() or None,
        "set_detail": (form.get("set_detail") or "").strip() or None,
        "detail": (form.get("detail") or "").strip() or None,
        "deadline": parse_date(form.get("deadline")) and parse_date(form.get("deadline")).isoformat(),
        "is_closed": form.get("is_closed") == "on",
        "note": (form.get("note") or "").strip() or None,
    }

# ----------------------------------------------------------------------------
# Admin — buyer orders + xlsx import
# ----------------------------------------------------------------------------

@app.get("/admin/orders", response_class=HTMLResponse)
async def admin_orders(request: Request, session_id: Optional[str] = Cookie(None), status: str = ""):
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    q = sb.table("go_buyer_orders").select("*").order("created_at", desc=True).limit(200)
    if status: q = q.eq("status", status)
    res = q.execute()
    orders = res.data or []
    # Fetch all items for these orders
    if orders:
        ids = [o["id"] for o in orders]
        items_res = sb.table("go_buyer_order_items").select("*").in_("order_id", ids).execute()
        by_order: Dict[int, List[Dict[str, Any]]] = {}
        for it in (items_res.data or []):
            by_order.setdefault(it["order_id"], []).append(it)
        for o in orders:
            o["items"] = by_order.get(o["id"], [])
    return templates.TemplateResponse("admin/orders.html", {"request": request, "orders": orders, "status": status})

@app.post("/admin/order/{order_id}/status")
async def admin_order_status(order_id: int, request: Request, session_id: Optional[str] = Cookie(None)):
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    form = await request.form()
    new_status = form.get("status")
    if new_status not in ("submitted", "confirmed", "shipped", "cancelled"):
        raise HTTPException(400, "Invalid status")
    sb.table("go_buyer_orders").update({"status": new_status, "updated_at": _now().isoformat()}).eq("id", order_id).execute()
    return RedirectResponse("/admin/orders", status_code=302)

@app.get("/admin/aggregate", response_class=HTMLResponse)
async def admin_aggregate(request: Request, session_id: Optional[str] = Cookie(None)):
    """Per-GO total quantity for Alibaba bulk-buy planning."""
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    res = sb.table("go_campaign_aggregate").select("*").order("status").order("deadline").execute()
    return templates.TemplateResponse("admin/aggregate.html", {"request": request, "rows": res.data or []})

@app.get("/admin/import", response_class=HTMLResponse)
async def admin_import_page(request: Request, session_id: Optional[str] = Cookie(None)):
    if (r := _admin_or_redirect(session_id)): return r
    return templates.TemplateResponse("admin/import.html", {"request": request, "result": None})

@app.post("/admin/import", response_class=HTMLResponse)
async def admin_import_submit(request: Request, session_id: Optional[str] = Cookie(None), file: UploadFile = File(...)):
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    import openpyxl
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    summary = {"POCA_SET": 0, "ALBUM": 0, "MD": 0, "errors": []}

    def upsert_rows(rows: List[Dict[str, Any]]):
        if not rows: return
        for row in rows:
            try:
                # Use go_code as conflict key
                existing = sb.table("go_campaigns").select("id").eq("go_code", row["go_code"]).execute()
                if existing.data:
                    sb.table("go_campaigns").update(row).eq("go_code", row["go_code"]).execute()
                else:
                    sb.table("go_campaigns").insert(row).execute()
            except Exception as e:
                summary["errors"].append(f"{row.get('go_code')}: {e}")

    # POCA SET
    if "POCA SET" in wb.sheetnames:
        ws = wb["POCA SET"]
        rows = []
        for r in ws.iter_rows(min_row=5, values_only=True):
            code = r[1]
            if not code or not str(code).startswith("GO-"): continue
            if not (r[2] and r[3]): continue  # need idol + album to be meaningful
            dl = parse_date(r[8])
            rows.append({
                "go_code": str(code).strip(), "category": "POCA_SET",
                "idol_name": str(r[2]).strip(), "album": str(r[3]).strip(),
                "retail_store": (str(r[4]).strip() if r[4] else None),
                "url": (str(r[5]).strip() if r[5] else None),
                "set_price_krw": parse_num(r[6]),
                "set_detail": (str(r[7]).strip() if r[7] else None),
                "deadline": dl.isoformat() if dl else None,
                "note": (str(r[11]).strip() if len(r) > 11 and r[11] else None),
            })
            summary["POCA_SET"] += 1
        upsert_rows(rows)

    # ALBUM
    if "ALBUM" in wb.sheetnames:
        ws = wb["ALBUM"]
        rows = []
        for r in ws.iter_rows(min_row=5, values_only=True):
            code = r[1]
            if not code or not str(code).startswith("ALBUM-"): continue
            if not (r[2] and r[3]): continue
            dl = parse_date(r[9])
            rows.append({
                "go_code": str(code).strip(), "category": "ALBUM",
                "idol_name": str(r[2]).strip(), "album": str(r[3]).strip(),
                "version": (str(r[4]).strip() if r[4] else None),
                "url": (str(r[5]).strip() if r[5] else None),
                "set_price_krw": parse_num(r[6]),
                "pob_store": (str(r[7]).strip() if r[7] else None),
                "pob_detail": (str(r[8]).strip() if r[8] else None),
                "deadline": dl.isoformat() if dl else None,
                "retail_price_krw": parse_num(r[12]) if len(r) > 12 else None,
                "discount_pct": parse_num(r[13]) if len(r) > 13 else None,
                "note": (str(r[14]).strip() if len(r) > 14 and r[14] else None),
            })
            summary["ALBUM"] += 1
        upsert_rows(rows)

    # MD (merchandise / 굿즈) — replaces former "ALBUM RESTOCK" sheet
    # New sheet name: "MD". Code prefix: MD-NNN.
    if "MD" in wb.sheetnames:
        ws = wb["MD"]
        rows = []
        for r in ws.iter_rows(min_row=5, values_only=True):
            code = r[1]
            if not code or not str(code).startswith("MD-"): continue
            if not (r[2] and r[3]): continue
            dl = parse_date(r[8]) if len(r) > 8 else None
            rows.append({
                "go_code": str(code).strip(), "category": "MD",
                "idol_name": str(r[2]).strip(), "album": str(r[3]).strip(),
                "retail_store": (str(r[4]).strip() if r[4] else None),
                "url": (str(r[5]).strip() if r[5] else None),
                "set_price_krw": parse_num(r[6]),
                "set_detail": (str(r[7]).strip() if r[7] else None),
                "deadline": dl.isoformat() if dl else None,
                "note": (str(r[11]).strip() if len(r) > 11 and r[11] else None),
            })
            summary["MD"] += 1
        upsert_rows(rows)

    return templates.TemplateResponse("admin/import.html", {"request": request, "result": summary})

# ----------------------------------------------------------------------------
# CSV exports — per-GO buyers + bulk for "what to bulk-buy on Alibaba this week"
# ----------------------------------------------------------------------------

# Convention: deadline = the listed date in KST until 23:59. After that, GO closes.

CSV_COLS = ["go_code", "order_id", "alibaba_username", "qty",
            "unit_price_krw", "line_total_krw", "contact_email", "contact_note",
            "created_at", "status"]

def _write_csv_rows(rows: List[Dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    buf.write("﻿")  # UTF-8 BOM so Excel opens Korean correctly
    w = csv.writer(buf)
    w.writerow(CSV_COLS)
    for r in rows:
        w.writerow([r.get(c, "") if r.get(c) is not None else "" for c in CSV_COLS])
    return buf.getvalue().encode("utf-8")

def _fetch_buyer_rows(go_codes: List[str]) -> List[Dict[str, Any]]:
    """Returns flattened rows joining go_buyer_orders + items, for the given GO codes,
    excluding cancelled orders. Sorted by created_at desc."""
    if not go_codes: return []
    items = sb.table("go_buyer_order_items").select("order_id, go_code, qty, unit_price_krw, line_total_krw").in_("go_code", go_codes).execute()
    item_rows = items.data or []
    if not item_rows: return []
    order_ids = list({i["order_id"] for i in item_rows})
    orders = sb.table("go_buyer_orders").select("id, alibaba_username, contact_email, contact_note, status, created_at").in_("id", order_ids).neq("status", "cancelled").execute()
    order_map = {o["id"]: o for o in (orders.data or [])}
    out: List[Dict[str, Any]] = []
    for it in item_rows:
        o = order_map.get(it["order_id"])
        if not o: continue  # cancelled or missing
        out.append({
            "go_code": it["go_code"], "order_id": it["order_id"],
            "alibaba_username": o["alibaba_username"], "qty": it["qty"],
            "unit_price_krw": it.get("unit_price_krw"), "line_total_krw": it.get("line_total_krw"),
            "contact_email": o.get("contact_email"), "contact_note": o.get("contact_note"),
            "created_at": (o.get("created_at") or "")[:19], "status": o["status"],
        })
    out.sort(key=lambda r: (r["go_code"], r.get("created_at") or ""))
    return out

@app.get("/admin/campaign/{code}/buyers.csv")
async def admin_export_buyers_csv(code: str, session_id: Optional[str] = Cookie(None)):
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    rows = _fetch_buyer_rows([code])
    body = _write_csv_rows(rows)
    fname = f"{code}_buyers_{date.today().isoformat()}.csv"
    return StreamingResponse(io.BytesIO(body), media_type="text/csv; charset=utf-8",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})

@app.get("/admin/buyers.csv")
async def admin_export_all_closing_csv(session_id: Optional[str] = Cookie(None), closing_within: int = 7):
    """One CSV with ALL buyers across GOs whose deadline is within N days. The 'what do I order this week' workflow."""
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return PlainTextResponse("DB not configured", status_code=503)
    today = date.today()
    cutoff = (today + timedelta(days=closing_within)).isoformat()
    res = sb.table("go_campaigns").select("go_code, deadline").eq("is_closed", False).not_.is_("deadline", "null").lte("deadline", cutoff).gte("deadline", today.isoformat()).execute()
    codes = [r["go_code"] for r in (res.data or [])]
    rows = _fetch_buyer_rows(codes)
    body = _write_csv_rows(rows)
    fname = f"GO_bulk_order_{today.isoformat()}_within{closing_within}d.csv"
    return StreamingResponse(io.BytesIO(body), media_type="text/csv; charset=utf-8",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})

# ----------------------------------------------------------------------------
# Cron: daily deadline check (D-day in {0,1,3,5,7}) — idempotent, rate-limited, heartbeat
# Set CRON_SECRET in Railway, then schedule via cron-job.org GET request.
# ----------------------------------------------------------------------------

_LAST_CRON_HIT: Dict[str, float] = {}
DDAY_THRESHOLDS = [0, 1, 3, 5, 7]

@app.get("/cron/deadline-check")
async def cron_deadline_check(secret: str = ""):
    expected = os.environ.get("CRON_SECRET", "")
    if not expected or secret != expected:
        raise HTTPException(401, "secret required")
    # Soft rate limit (per process): 1 call / 60s. Returns 200+skipped instead of 429 so cron-job.org's
    # retry-on-failure logic doesn't loop or send false-failure emails.
    nowt = time.time()
    if nowt - _LAST_CRON_HIT.get("deadline-check", 0) < 60:
        return {"ok": True, "skipped": "rate-limited (1 call / 60s)", "scope": f"deadline-check:{date.today().isoformat()}"}
    _LAST_CRON_HIT["deadline-check"] = nowt
    if not sb:
        return {"ok": False, "error": "db not configured"}

    today = date.today()
    scope = f"deadline-check:{today.isoformat()}"
    # Idempotency: if a row for (job, scope) already exists, this is a retry — skip Telegram send
    prev = sb.table("cron_runs").select("id, ran_at").eq("job", "deadline-check").eq("scope", scope).execute()
    if prev.data:
        return {"ok": True, "skipped": "already ran today", "scope": scope, "first_ran_at": prev.data[0]["ran_at"]}

    # Pull all open campaigns with a deadline, group by D-day
    res = sb.table("go_campaigns").select("go_code, idol_name, album, retail_store, version, deadline").eq("is_closed", False).not_.is_("deadline", "null").execute()
    by_dday: Dict[int, List[Dict[str, Any]]] = {k: [] for k in DDAY_THRESHOLDS}
    for r in (res.data or []):
        d = parse_date(r.get("deadline"))
        if not d: continue
        delta = (d - today).days
        if delta in by_dday:
            by_dday[delta].append(r)

    # Build digest (always send — heartbeat ensures silence ≠ healthy)
    lines = ["📅 GO 마감 알림 (자동)"]
    has_any = any(by_dday.values())
    if has_any:
        for k, label in [(0, "🔴 D-0 (오늘 마감 — 23:59 KST)"), (1, "🚨 D-1"), (3, "⚠ D-3"), (5, "D-5"), (7, "D-7")]:
            if by_dday[k]:
                lines.append(f"\n{label}:")
                for it in by_dday[k]:
                    code = it["go_code"]; idol = it["idol_name"]; alb = it["album"]
                    extra = it.get("retail_store") or it.get("version") or ""
                    lines.append(f"  • {code} {idol} {alb} {extra}".rstrip())
    else:
        lines.append("\n오늘 마감 임박(0/1/3/5/7일) 항목 없음 (heartbeat ✓)")
    lines.append(f"\n— {today.isoformat()} 09:00 KST 기준 / starphotocard-go")
    msg = "\n".join(lines)
    sent = await notify_admin(msg, link_url="https://starphotocard-go.up.railway.app/admin/aggregate")

    # Audit (records dedup key — second call returns "skipped")
    try:
        sb.table("cron_runs").insert({
            "job": "deadline-check", "scope": scope,
            "result": {"sent": sent, "by_dday": {str(k): [r["go_code"] for r in v] for k, v in by_dday.items()}},
        }).execute()
    except Exception as e:
        print(f"[cron_deadline_check] audit insert failed: {e}")
    return {"ok": True, "sent": sent, "scope": scope, "by_dday": {k: len(v) for k, v in by_dday.items()}}

# ----------------------------------------------------------------------------
# FAQ (public)
# ----------------------------------------------------------------------------

@app.get("/faq", response_class=HTMLResponse)
async def faq_page(request: Request):
    return templates.TemplateResponse("faq.html", {"request": request})

# ----------------------------------------------------------------------------
# Kakao OAuth (admin-only) — connect once to authorize, then notifications work.
# ----------------------------------------------------------------------------

@app.get("/admin/kakao/connect", response_class=HTMLResponse)
async def admin_kakao_connect(session_id: Optional[str] = Cookie(None)):
    if (r := _admin_or_redirect(session_id)): return r
    if not KAKAO_REST_API_KEY:
        return PlainTextResponse("KAKAO_REST_API_KEY not configured in Railway env", status_code=503)
    auth_url = (f"{KAKAO_AUTH_URL}?response_type=code"
                f"&client_id={KAKAO_REST_API_KEY}"
                f"&redirect_uri={KAKAO_REDIRECT_URI}"
                f"&scope=talk_message")
    return RedirectResponse(auth_url, status_code=302)

@app.get("/admin/kakao/callback", response_class=HTMLResponse)
async def admin_kakao_callback(request: Request, code: str = "", error: str = "", session_id: Optional[str] = Cookie(None)):
    if (r := _admin_or_redirect(session_id)): return r
    if error:
        return PlainTextResponse(f"Kakao authorization denied: {error}", status_code=400)
    if not code:
        return PlainTextResponse("Missing ?code from Kakao callback", status_code=400)
    if not KAKAO_REST_API_KEY:
        return PlainTextResponse("KAKAO_REST_API_KEY not configured", status_code=503)
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            data = {
                "grant_type": "authorization_code",
                "client_id": KAKAO_REST_API_KEY,
                "redirect_uri": KAKAO_REDIRECT_URI,
                "code": code,
            }
            if KAKAO_CLIENT_SECRET:
                data["client_secret"] = KAKAO_CLIENT_SECRET
            r = await client.post(KAKAO_TOKEN_URL, data=data)
            j = r.json()
            if r.status_code != 200 or not j.get("access_token"):
                return PlainTextResponse(f"Kakao token exchange failed: {j}", status_code=502)
            await _kakao_save_token(
                access_token=j["access_token"],
                refresh_token=j.get("refresh_token"),
                expires_in=j.get("expires_in"),
                refresh_expires_in=j.get("refresh_token_expires_in"),
                scope=j.get("scope"),
            )
        return HTMLResponse(
            "<h1>✅ Kakao connected</h1>"
            "<p>알림이 이제부터 본인 카카오톡으로 발송됩니다.</p>"
            "<p>다음 단계: <a href='/admin/kakao/test'>📨 테스트 메시지 발송</a></p>"
            "<p><a href='/admin/dashboard'>← Admin Dashboard</a></p>"
        )
    except Exception as e:
        return PlainTextResponse(f"Kakao callback exception: {e}", status_code=500)

@app.get("/admin/kakao/test")
async def admin_kakao_test(session_id: Optional[str] = Cookie(None)):
    """Send a test "memo to self" with full error diagnostics so we don't have to read Railway logs."""
    if (r := _admin_or_redirect(session_id)): return r
    diagnostic: Dict[str, Any] = {"step": "start"}
    try:
        diagnostic["step"] = "load_token"
        row = await _kakao_load_token()
        diagnostic["token_row_present"] = bool(row)
        diagnostic["has_access"] = bool(row and row.get("access_token"))
        diagnostic["has_refresh"] = bool(row and row.get("refresh_token"))
        diagnostic["scope"] = row.get("scope") if row else None
        diagnostic["access_expires_at"] = row.get("access_expires_at") if row else None

        diagnostic["step"] = "get_valid_access"
        access = await _kakao_get_valid_access_token()
        diagnostic["got_access_token"] = bool(access)
        if not access:
            return JSONResponse({"ok": False, "stage": "no_access_token", "diagnostic": diagnostic,
                                 "hint": "Click /admin/kakao/connect again to re-authorize."}, status_code=502)

        diagnostic["step"] = "send_memo"
        # Minimal text template — link required by Kakao API.
        template_object = {
            "object_type": "text",
            "text": "🔧 starphotocard-go test message — Kakao 연동 확인용입니다. 실제 운영 시에는 새 주문이 들어오면 자동 알림이 옵니다.",
            "link": {"web_url": "https://starphotocard-go.up.railway.app",
                     "mobile_web_url": "https://starphotocard-go.up.railway.app"},
        }
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.post(KAKAO_SEND_URL,
                                  headers={"Authorization": f"Bearer {access}"},
                                  data={"template_object": json_dumps(template_object)})
            diagnostic["kakao_status"] = r.status_code
            try:
                diagnostic["kakao_body"] = r.json()
            except Exception:
                diagnostic["kakao_body"] = r.text[:400]
            if r.status_code == 200:
                return HTMLResponse(
                    "<h1>✅ 테스트 메시지 발송 완료</h1>"
                    "<p>본인 카카오톡 채팅 목록에서 '나에게 보내기' 방을 확인해주세요.</p>"
                    "<p><a href='/admin/dashboard'>← Admin Dashboard</a></p>"
                    f"<details><summary>diagnostic</summary><pre>{json_dumps(diagnostic)}</pre></details>"
                )
            return JSONResponse({"ok": False, "stage": "kakao_api_error", "diagnostic": diagnostic}, status_code=502)
    except Exception as e:
        diagnostic["exception"] = f"{type(e).__name__}: {e}"
        return JSONResponse({"ok": False, "stage": "exception", "diagnostic": diagnostic}, status_code=500)

@app.get("/admin/kakao/status")
async def admin_kakao_status(session_id: Optional[str] = Cookie(None)):
    if (r := _admin_or_redirect(session_id)): return r
    if not sb: return JSONResponse({"ok": False, "error": "db not configured"}, status_code=503)
    row = await _kakao_load_token()
    if not row:
        return JSONResponse({"ok": False, "connected": False, "reason": "no token row"})
    has_access = bool(row.get("access_token"))
    has_refresh = bool(row.get("refresh_token"))
    return JSONResponse({
        "ok": True, "connected": has_access and has_refresh,
        "has_access_token": has_access, "has_refresh_token": has_refresh,
        "access_expires_at": row.get("access_expires_at"),
        "refresh_expires_at": row.get("refresh_expires_at"),
        "scope": row.get("scope"), "updated_at": row.get("updated_at"),
        "rest_api_key_configured": bool(KAKAO_REST_API_KEY),
    })
